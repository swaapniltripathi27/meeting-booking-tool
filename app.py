from flask import Flask, render_template, request, jsonify, redirect
from openpyxl import Workbook, load_workbook
import os
import sqlite3
import smtplib
from email.message import EmailMessage
from ics import Calendar, Event

app = Flask(__name__)


# ----------------------------
# EMAIL CONFIGURATION
# ----------------------------

SENDER_EMAIL = "swaapnil.tripathi27@gmail.com"
SENDER_PASSWORD = "vgkzsstwnoyfzeqd"


# ----------------------------
# TECHM LEADER LIST
# ----------------------------

leaders = [
    "Mohit Joshi",
    "Manish Mangal",
    "Harshul Asnani",
    "Jena",
    "Sandeep",
    "Harsh",
    "Dom",
    "Gaurav",
    "Praveen",
    "Ashwini",
    "Swaapnil",
    "Arun",
    "Vasu",
    "Hridey"
]


# ----------------------------
# LEADER EMAIL MAPPING
# ----------------------------

leader_emails = {
    "Mohit Joshi": "mohit.joshi@techmahindra.com",
    "Manish Mangal": "manish.mangal@techmahindra.com",
    "Harshul Asnani": "harshul.asnani@techmahindra.com",
    "Jena": "jena@techmahindra.com",
    "Sandeep": "sandeep@techmahindra.com",
    "Harsh": "harsh@techmahindra.com",
    "Dom": "dom@techmahindra.com",
    "Gaurav": "gaurav@techmahindra.com",
    "Praveen": "praveen@techmahindra.com",
    "Ashwini": "ashwinikarthik.jonnakadla@TechMahindra.com",
    "Swaapnil": "swaapnil.tripathi@TechMahindra.com",
    "Arun": "AR0060126@TechMahindra.com",
    "Vasu": "Vasu.Mudgal@TechMahindra.com",
    "Hridey":"Hridey.Asnani@TechMahindra.com"
}

# ----------------------------
# DATABASE INITIALIZATION
# ----------------------------

def init_db():

    conn = sqlite3.connect("database.db")
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS meetings(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        time TEXT,
        leader TEXT,
        client_name TEXT,
        designation TEXT,
        organisation TEXT,
        opportunity_size TEXT
    )
    """)

    conn.commit()
    conn.close()


# ----------------------------
# CALENDAR INVITE FUNCTION
# ----------------------------
from datetime import datetime


def send_calendar_invite(leader, date, time, client, organisation):

    leader_email = leader_emails.get(leader)

    if not leader_email:
        return

    start_time = time.split("-")[0]
    end_time = time.split("-")[1]

    # Convert date like "2 March" to proper format
    year = 2026
    date_obj = datetime.strptime(f"{date} {year}", "%d %B %Y")

    start_datetime = datetime.strptime(
        f"{date_obj.date()} {start_time}", "%Y-%m-%d %H:%M"
    )

    end_datetime = datetime.strptime(
        f"{date_obj.date()} {end_time}", "%Y-%m-%d %H:%M"
    )

    event = Event()
    event.name = f"Meeting with {organisation}"
    event.begin = start_datetime
    event.end = end_datetime
    event.description = f"Client: {client}\nOrganisation: {organisation}"

    calendar = Calendar()
    calendar.events.add(event)

    msg = EmailMessage()
    msg["Subject"] = f"Meeting Invite - {organisation}"
    msg["From"] = SENDER_EMAIL
    msg["To"] = leader_email

    msg.set_content("Calendar invite attached")

    msg.add_attachment(
        str(calendar),
        subtype="calendar",
        filename="meeting_invite.ics"
    )

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(SENDER_EMAIL, SENDER_PASSWORD)
        smtp.send_message(msg)


# ----------------------------
# HOME PAGE
# ----------------------------

@app.route("/")
def home():

    dates = ["2 March", "3 March", "4 March"]

    times = []

    for hour in range(9, 18):
        times.append(f"{hour}:00-{hour}:30")
        times.append(f"{hour}:30-{hour+1}:00")

    return render_template(
        "index.html",
        dates=dates,
        times=times,
        leaders=leaders
    )


# ----------------------------
# CHECK AVAILABLE LEADERS
# ----------------------------

@app.route("/available_leaders")
def available_leaders():

    date = request.args.get("date")
    time = request.args.get("time")

    conn = sqlite3.connect("database.db")
    c = conn.cursor()

    c.execute(
        "SELECT leader FROM meetings WHERE date=? AND time=?",
        (date, time)
    )

    booked = [row[0] for row in c.fetchall()]

    conn.close()

    available = [l for l in leaders if l not in booked]

    return jsonify({"leaders": available})


# ----------------------------
# BOOK MEETING
# ----------------------------

@app.route("/book", methods=["POST"])
def book():

    date = request.form["date"]
    time = request.form["time"]
    leader = request.form["leader"]
    client = request.form["client"]
    designation = request.form["designation"]
    organisation = request.form["organisation"]
    opportunity = request.form["opportunity"]

    conn = sqlite3.connect("database.db")
    c = conn.cursor()

    c.execute("""
    INSERT INTO meetings
    (date, time, leader, client_name, designation, organisation, opportunity_size)
    VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (date, time, leader, client, designation, organisation, opportunity))

    conn.commit()
    conn.close()
    save_to_excel(date, time, leader, client, designation, organisation, opportunity)

    # SEND OUTLOOK CALENDAR INVITE
    send_calendar_invite(leader, date, time, client, organisation)

    return redirect("/")

def save_to_excel(date, time, leader, client, designation, organisation, opportunity):

    file = "meetings.xlsx"

    # Create file if it does not exist
    if not os.path.exists(file):

        wb = Workbook()
        ws = wb.active

        ws.append([
            "Date",
            "Time",
            "Leader",
            "Client Name",
            "Designation",
            "Organisation",
            "Opportunity Size"
        ])

        wb.save(file)

    wb = load_workbook(file)
    ws = wb.active

    ws.append([
        date,
        time,
        leader,
        client,
        designation,
        organisation,
        opportunity
    ])

    wb.save(file)


# ----------------------------
# START APP
# ----------------------------

if __name__ == "__main__":

    init_db()
    app.run(debug=True, port=8000)